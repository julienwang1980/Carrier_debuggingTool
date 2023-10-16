# -*- coding: utf-8 -*-

import sys
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
# import Ui_MainDesk
from Ui_Serial_connect import Ui_Serial_connect
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter



class MainDesk(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainDesk, self).__init__()
        self.setupUi(self)
        # 菜单的点击事件，当点击关闭菜单时连接槽函数 close()
        # self.actionExit.triggered.connect(self.close)
        # 菜单的点击事件，当点击打开菜单时连接槽函数 openMsg()










if __name__ == '__main__':
    # 创建QApplication类的实例
    app = QApplication(sys.argv)
    # 创建一个窗口
    MainDesk = MainDesk()
    # 显示窗口
    MainDesk.show()
    sys.exit(app.exec_())

# wb = load_workbook('./data.xlsx')
# ws = wb['Sheet1']
# data = []
# for row in ws.iter_rows(min_row=3, max_row=3, min_col=2):
#     for cell in row:
#         data.append(cell.value)
# wb.close()
# print(data)
# print("on_connet")