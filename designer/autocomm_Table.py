# -*- coding: utf-8 -*-

import sys
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
import modbus_tk.defines as cst
from PyQt5.QtCore import QTimer
import struct
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Alignment, Border, Side
import time


class AutocommTableSubwindow(QMdiSubWindow, ):
    modbus = None
    scanRate = 5000
    widgetSize = {"width": 250, "height": 150}

    err_code = {'100101': '配置率超上限', '100102': '配置率超下限', '100201': '环境温度超上限', '100202': '环境温度超下限',
                '100301': '外机风机异常', '100401': '压缩机启动异常', '100501': '手阀未开启', '100601': '主阀动作异常',
                '100701': '四通阀动作异常', '200101': '内机EXV异常', '200102': '内机风机异常', '200103': '内机风机过半异常',
                '300101': '系统高压报警', '300102': '系统低压报警', '300103': '排气温度过高', '300104': '驱动器温度过高'}

    def __init__(self):
        super(AutocommTableSubwindow, self).__init__()
        self.resize(self.widgetSize["width"], self.widgetSize["height"])
        self.SubWinWidget = QWidget()
        self.setWidget(self.SubWinWidget)
        layout = QVBoxLayout(self.SubWinWidget)
        self.tablewidget = QTableWidget()
        self.tablewidget.setRowCount(1)
        self.tablewidget.setColumnCount(1)
        layout.addWidget(self.tablewidget)
        self.setWindowTitle("autocommisioning")
        self.tablewidget.setHorizontalHeaderLabels(['autocommi_timer'])
        self.tablewidget.resizeColumnsToContents()
        self.tablewidget.resizeRowsToContents()
        self.setLayout(layout)
        self.mTimer = QTimer()
        self.mTimer.timeout.connect(self.ReceiverPortData)
        self.mTimer.start(self.scanRate)
        self.odus = []
        self.idus = []
        # 读取表格模板
        self.wb = load_workbook("../多联机调试记录表-自动调试.xlsx")

        # f = 0
        # z0=hex(65535)[2:]
        # z1=hex(65535)[2:]
        # z = z0+z1
        # print(z0+z1)
        # f = struct.unpack('!f', bytes.fromhex(z))[0]
        # print(f)

        # self.wb = load_workbook("../多联机调试记录表-自动调试.xlsx")
        #
        # self.wb.save("../多联机调试记录表_temp.xlsx")

        # col = column_index_from_string('bb')

    def closeEvent(self, event):
        # 获取主窗口
        parent = self.parent().parent().parent()
        parent.AutomissionStart.setEnabled(True)
        # print("程序执行到文件：{}，第{}行".format(sys._getframe().f_code.co_filename, str(sys._getframe().f_lineno)))
        # 表格<故障表>数据
        line_num = 0
        for i in self.odus:
            if i["err_code"] != 0:
                error_line = ["{}#外机".format(i['addr'])]
                error_line.append(i["err_code"])
                error_line.append(self.err_code[str(i["err_code"])])
                self.wb["故障表"].append(error_line)
                line_num += 1
        for i in self.idus:
            if i["err_code"] != 0:
                error_line = ["{}#内机".format(i['addr'])]
                error_line.append(i["err_code"])
                error_line.append(self.err_code[str(i["err_code"])])
                self.wb["故障表"].append(error_line)
                line_num += 1
        thin = Side(border_style="thin", color="000000")
        for row in self.wb["故障表"].iter_rows(min_row=3, max_col=3, max_row=line_num + 2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False,
                                           shrink_to_fit=False)  # 对齐样式
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)  # 边框

        # 保存并关闭表格
        localtime = time.strftime("%Y%m%d%H%M", time.localtime())
        self.wb.save("../多联机调试记录表_{}.xlsx".format(localtime))
        self.wb.close()
        self.deleteLater()

    def ReceiverPortData(self):
        # print("程序执行到文件：{}，第{}行".format(sys._getframe().f_code.co_filename, str(sys._getframe().f_lineno)))
        # print(self.wb.sheetnames)
        if self.modbus:
            try:
                # Connect to the slave
                self.modbus['master'].set_timeout(1)
                self.modbus['master'].set_verbose(True)
                data = []
                for i in range(8):
                    # 读取下位机1的寄存器值，采用cst.READ_INPUT_REGISTERS，地址40000开始，长度100，数据格式为short
                    x = self.modbus['master'].execute(1, cst.READ_INPUT_REGISTERS, 40000 + 100 * i, 100,
                                                      data_format=">{}".format(100 * 'h'))
                    data.extend(x)
                    print(x)
                newItem = QTableWidgetItem(str(data[0]))
                self.tablewidget.setItem(0, 0, newItem)
                # 写入excel表内
                if len(self.odus) == 0:
                    if data[0] >= 360 - 10:
                        for i in range(20, 100, 20):
                            if data[i] != 0:
                                odu = {}
                                odu['addr'] = int(i / 20 - 1)
                                odu['HP'] = data[i]
                                odu['err_code'] = 0
                                self.odus.append(odu)
                        for i in range(100, 800, 7):
                            if (data[i] != 0 or data[i + 1] != 0 or data[i + 2] != 0):
                                idu = {}
                                idu['addr'] = int((i - 100) / 7)
                                idu['err_code'] = 0
                                self.idus.append(idu)
                        align = Alignment(horizontal='center', vertical='center', wrap_text=False,
                                          shrink_to_fit=False)  # 居中样式
                        # 表格<详细数据-外机>表头
                        row = self.wb["详细数据-外机"]['2']
                        for i in range(1, len(self.odus)):
                            col_index = 2 + 12 * i
                            col_lett = get_column_letter(col_index)
                            self.wb["详细数据-外机"]['{}1'.format(col_lett)] = "{}#外机".format(self.odus[i]["addr"])
                            self.wb["详细数据-外机"]['{}1'.format(col_lett)].alignment = align
                            self.wb["详细数据-外机"].merge_cells(start_row=1, start_column=col_index, end_row=1,
                                                           end_column=col_index + 11)
                            for j in range(12):
                                col_lett = get_column_letter(col_index + j)
                                self.wb["详细数据-外机"]['{}2'.format(col_lett)] = row[j + 1].value
                                self.wb["详细数据-外机"]['{}2'.format(col_lett)].alignment = align
                        # 表格<详细数据-内机>表头
                        row = self.wb["详细数据-内机"]['2']
                        for i in range(len(self.idus)):
                            col_index = 2 + 6 * i
                            col_lett = get_column_letter(col_index)
                            self.wb["详细数据-内机"]['{}1'.format(col_lett)] = "{}#内机".format(self.idus[i]["addr"])
                            self.wb["详细数据-内机"]['{}1'.format(col_lett)].alignment = align
                            self.wb["详细数据-内机"].merge_cells(start_row=1, start_column=col_index, end_row=1,
                                                           end_column=col_index + 5)
                            for j in range(6):
                                col_lett = get_column_letter(col_index + j)
                                self.wb["详细数据-内机"]['{}2'.format(col_lett)] = row[j + 1].value
                                self.wb["详细数据-内机"]['{}2'.format(col_lett)].alignment = align
                else:
                    # 表格<详细数据-外机>数据
                    line_odu = [data[0]]
                    for i in range(len(self.odus)):
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 1] / 10)
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 2] / 10)
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 3] / 10)
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 4] / 10)
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 5])
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 6])
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 7])
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 8])
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 9])
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 10])
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 11])
                        line_odu.append(data[20 + 20 * self.odus[i]['addr'] + 12])
                        self.odus[i]["err_code"] = data[20 + 20 * i + 13] // 1000 * 100000 + data[
                            20 + 20 * i + 13] % 1000
                    self.wb["详细数据-外机"].append(line_odu)
                    # 表格<详细数据-内机>数据
                    line_idu = [data[0]]
                    for i in range(len(self.idus)):
                        line_idu.append(data[100 + 7 * self.idus[i]['addr']] / 10)
                        line_idu.append(data[100 + 7 * self.idus[i]['addr'] + 1] / 10)
                        line_idu.append(data[100 + 7 * self.idus[i]['addr'] + 2] / 10)
                        line_idu.append(data[100 + 7 * self.idus[i]['addr'] + 3])
                        line_idu.append(data[100 + 7 * self.idus[i]['addr'] + 4])
                        line_idu.append(data[100 + 7 * self.idus[i]['addr'] + 5])
                        self.idus[i]["err_code"] = data[100 + 7 * self.idus[i]['addr'] + 6] // 1000 * 100000 + data[
                            100 + 7 * self.idus[i]['addr'] + 6] % 1000
                    self.wb["详细数据-内机"].append(line_idu)
            except:
                print('unknown')
                print("run")

        # data1 = []
        # data1.append(-1)
        # newItem = QTableWidgetItem(str(data1[0]))
        # self.tablewidget.setItem(0, 0, newItem)
        # self.wb["调试记录表"]
        # data = []
        # for row in ws.iter_rows(min_row=3, max_row=3, min_col=2):
        #     for cell in row:
        #         data.append(cell.value)
        # wb.close()
        # self.table1.modbus = self.modbus_conn


if __name__ == '__main__':
    # 创建QApplication类的实例
    app = QApplication(sys.argv)
    # 创建一个窗口
    MainDesk = AutocommTableSubwindow()
    # 显示窗口
    MainDesk.show()
    sys.exit(app.exec_())
